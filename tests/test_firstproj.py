import openpyxl
from openpyxl import load_workbook

import main


def test_get_data():
    result = main.get_data(url="https://api.data.gov/ed/collegescorecard/v1/schools.json?"
                               "school.degrees_awarded.predominant=2,"
                               "3&fields=id,school.state,school.city,school.name,2018.student.size,"
                               "2016.repayment.3_yr_repayment.overall,"
                               "2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line,2017.student.size,"
                               "2016.repayment.repayment_cohort.3_year_declining_balance")
    count = 0
    for element in result:
        count += 1
        return count

    assert count > 1000


def test_write_to_table_one():
    conn, cursor = main.open_db("db_test_write_one.sqlite")
    print(type(conn))
    main.setup_db(cursor)

    all_data = [{'2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line': 187,
                 'school.name': 'Kauai Community College', '2017.student.size': 1007, '2018.student.size':
                     929, 'school.state': 'HI', 'id': 141802, 'school.city': 'Lihue',
                 '2016.repayment.3_yr_repayment.overall': 273,
                 '2016.repayment.repayment_cohort.3_year_declining_balance': 0.5018315018}]

    for item in all_data:
        main.make_database_data(cursor, item['school.name'], item['school.city'], item['school.state'],
                                item['2018.student.size'],
                                item['2017.student.size'],
                                item['2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line'],
                                item['2016.repayment.3_yr_repayment.overall'],
                                item['2016.repayment.repayment_cohort.3_year_declining_balance'])

    row_count = cursor.rowcount

    assert row_count == 1

    main.close_db(conn)


def test_write_to_table_two():
    conn, cursor = main.open_db("db_test_write_two.sqlite")
    print(type(conn))
    main.setup_db(cursor)

    workbook = load_workbook(filename="state_M2019_dl.xlsx")
    sheet = workbook.active

    for value in sheet.iter_rows(values_only=True):
        main.make_wage_database_data(cursor, value[1], value[9], value[8], value[10], value[19], value[7])

    row_count = cursor.rowcount

    assert row_count == 1

    main.close_db(conn)


def test_table_one():
    conn, cursor = main.open_db("db_test_table1.sqlite")
    print(type(conn))
    main.setup_db(cursor)
    all_data = []

    for item in all_data:
        main.make_database_data(cursor, "test_school",
                                "test_city",
                                "test_state",
                                "test_2018_size",
                                "test_2017_size",
                                "test_2017_poverty",
                                "test_2016_repayment",
                                "test_2016_repayment_cohort")
        assert all_data[0] == "test_school"

    main.close_db(conn)


def test_table_two():
    conn, cursor = main.open_db("db_test_table2.sqlite")
    print(type(conn))
    main.setup_db(cursor)
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for value in sheet.iter_rows(values_only=True):
        main.make_wage_database_data(cursor, "test1", "test2", "test3", "test4", "test5")

        assert value[1] == "test1"

    main.close_db(conn)


def test_data_reload():
    new_data_api = main.get_api_data()
    new_data_excel = main.get_excel_data()

    assert new_data_api[0] == {'name': 'Kauai Community College', 'city': 'Lihue', 'cstate': 'HI', 'size_2018': 929,
                               'size_2017': 1007, 'poverty_2017': 187, 'repayment_2016': 273,
                               'repayment_cohort_2016': 0.5018315018}

    assert new_data_excel[0] == {'state': 'Alabama', 'occ_group': 'major', 'major_title': 'Management Occupations',
                                 'total_employment': 83760, 'percentile_25_salary': 31.8, 'occupation_code': '11-0000'}
