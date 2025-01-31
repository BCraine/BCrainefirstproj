# Brandon Craine
# import openpyxl
import requests
from openpyxl import load_workbook
import PySide6
import sys
import pandas
import plotly

import GuiWindow
import secrets
import sqlite3
from typing import Tuple, List, Dict

statedict = {}


def display_data(data: list, data1: list):
    qt_app = PySide6.QtWidgets.QApplication(sys.argv)
    my_window = GuiWindow.BCraineGuiWindow(data, data1)
    my_window
    sys.exit(qt_app.exec_())


def get_data(url: str):
    all_data = []

    for page in range(0, 161):
        full_url = f"{url}&api_key={secrets.api_key}&page={page}"

        response = requests.get(full_url)

        if response.status_code != 200:
            print(response.text)
            return []
        json_data = response.json()

        results = json_data['results']

        all_data.extend(results)

    return all_data


def open_db(filename: str) -> Tuple[sqlite3.Connection, sqlite3.Cursor]:
    db_connection = sqlite3.connect(filename)  # connect to existing DB or create new one

    cursor = db_connection.cursor()  # get ready to read/write data

    return db_connection, cursor


def close_db(connection: sqlite3.Connection):
    connection.commit()  # make sure any changes get saved
    connection.close()


def setup_db(cursor: sqlite3.Cursor):
    cursor.execute('''CREATE TABLE IF NOT EXISTS college(
    college_name TEXT NOT NULL,
    college_city TEXT NOT NULL,
    college_state TEXT NOT NULL,
    student_size_2018 INTEGER DEFAULT 0,
    student_size_2017 INTEGER DEFAULT 0,
    earnings_3yrs_after_completion_overall_count_over_poverty_line_2017 INTEGER DEFAULT 0,
    repayment_3_yr_repayment_overall_2016 INTEGER DEFAULT 0,
    repayment_repayment_cohort_3_year_declining_balance_2016 REAL DEFAULT 0,
    PRIMARY KEY(college_name, college_city, college_state)
    );''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS wage_data(
    data_state TEXT NOT NULL,
    occ_group TEXT NOT NULL,
    major_title TEXT NOT NULL,
    total_employment_field_in_state INTEGER DEFAULT 0,
    percentile_25_salary REAL DEFAULT 0,
    occupation_code REAL DEFAULT 0);''')


def make_database_data(cursor: sqlite3.Cursor, name, city, state, size_2018, size_2017, earnings_2017, repayment_2016,
                       repayment_cohort_2016):
    if earnings_2017 is None:
        earnings_2017 = 0

    if repayment_2016 is None:
        repayment_2016 = 0

    if size_2017 is None:
        size_2017 = 0

    if size_2018 is None:
        size_2018 = 0

    if repayment_cohort_2016 is None:
        repayment_cohort_2016 = 0

    cursor.execute('''INSERT INTO college (college_name,college_city,college_state,
    student_size_2018,student_size_2017,earnings_3yrs_after_completion_overall_count_over_poverty_line_2017,
    repayment_3_yr_repayment_overall_2016, repayment_repayment_cohort_3_year_declining_balance_2016)
    VALUES(?,?,?,?,?,?,?,?)''', (name, city, state, size_2018, size_2017, earnings_2017, repayment_2016,
                                 repayment_cohort_2016))


def make_wage_database_data(cursor: sqlite3.Cursor, d_state, o_group, m_title, t_employment_f_state, p_25_salary,
                            o_code):
    if o_group == "major":
        cursor.execute('''INSERT INTO wage_data(data_state, occ_group, major_title,total_employment_field_in_state,
        percentile_25_salary,
                                occupation_code) values (?,?,?,?,?,?)''', (d_state, o_group, m_title,
                                                                           t_employment_f_state,
                                                                           p_25_salary, o_code))


# def get_excel_data() -> List[Dict]:
# workbook_file = openpyxl.load_workbook("state_M2019_dl.xlsx")
# sheet = workbook_file.active
# final_data_list = []
# for value in sheet.iter_rows(values_only=True):
#  record = {"state": value[1], "occ_group": value[9], "major_title": value[8], "total_employment": value[10],
#  "percentile_25_salary": value[19], "occupation_code": value[7]}

# final_data_list.append(record)
# return final_data_list
def get_api_data() -> List[Dict]:
    final_data_list_table_one = []
    url = "https://api.data.gov/ed/collegescorecard/v1/schools.json?school.degrees_awarded.predominant=2," \
          "3&fields=id,school.state,school.city,school.name,2018.student.size," \
          "2016.repayment.3_yr_repayment.overall," \
          "2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line,2017.student.size," \
          "2016.repayment.repayment_cohort.3_year_declining_balance"

    all_data = get_data(url)
    for item in all_data:
        record1 = {"name": item['school.name'], "city": item['school.city'], "cstate": item['school.state'],
                   "size_2018": item['2018.student.size'],
                   "size_2017": item['2017.student.size'],
                   "poverty_2017": item['2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line'],
                   "repayment_2016": item['2016.repayment.3_yr_repayment.overall'],
                   "repayment_cohort_2016": item['2016.repayment.repayment_cohort.3_year_declining_balance']}
        final_data_list_table_one.append(record1)
    return final_data_list_table_one


def get_excel_data() -> List[Dict]:
    workbook = load_workbook(filename="state_M2019_dl.xlsx")
    sheet = workbook.active
    final_data_list_table_two = []
    for value in sheet.iter_rows(values_only=True):

        if value[9] == "major":
            record = {"state": value[1], "occ_group": value[9], "major_title": value[8], "total_employment": value[10],
                      "percentile_25_salary": value[19], "occupation_code": value[7]}

            final_data_list_table_two.append(record)
    return final_data_list_table_two


def get_key(value: dict):
    return value["occupation_code"]


def excel_ascend_function():
    act_2 = get_excel_data()
    act_2.sort(key=get_key)

    return act_2


def reload_data():
    display_data(get_api_data(), excel_ascend_function())


def main():
    # try:

    conn, cursor = open_db("bcrainedb.sqlite")
    print(type(conn))
    setup_db(cursor)
    final_data_list_table_one = []

    url = "https://api.data.gov/ed/collegescorecard/v1/schools.json?school.degrees_awarded.predominant=2," \
          "3&fields=id,school.state,school.city,school.name,2018.student.size," \
          "2016.repayment.3_yr_repayment.overall," \
          "2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line,2017.student.size," \
          "2016.repayment.repayment_cohort.3_year_declining_balance"

    all_data = get_data(url)

    outfile = open('schooldata.txt', 'w')
    datastring = ','.join([str(i) for i in all_data])
    outfile.write(datastring)
    outfile.close()

    for item in all_data:
        make_database_data(cursor, item['school.name'], item['school.city'], item['school.state'],
                           item['2018.student.size'],
                           item['2017.student.size'],
                           item['2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line'],
                           item['2016.repayment.3_yr_repayment.overall'],
                           item['2016.repayment.repayment_cohort.3_year_declining_balance'])

        record1 = {"name": item['school.name'], "city": item['school.city'], "cstate": item['school.state'],
                   "size_2018": item['2018.student.size'],
                   "size_2017": item['2017.student.size'],
                   "poverty_2017": item['2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line'],
                   "repayment_2016": item['2016.repayment.3_yr_repayment.overall'],
                   "repayment_cohort_2016": item['2016.repayment.repayment_cohort.3_year_declining_balance']}
        final_data_list_table_one.append(record1)

    workbook = load_workbook(filename="state_M2019_dl.xlsx")
    sheet = workbook.active
    states = ["Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado", "Connecticut", "Delaware",
              "District of Columbia", "Florida", "Georgia", "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa", "Kansas",
              "Kentucky", "Louisiana", "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota", "Mississippi",
              "Missouri", "Montana", "Nebraska", "Nevada", "New Hampshire", "New Jersey", "New Mexico", "New York",
              "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon", "Pennsylvania", "Rhode Island",
              "South Carolina", "South Dakota", "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
              "West Virginia", "Wisconsin", "Wyoming", "Guam", "Puerto Rico", "Virgin Islands"]
    # for value1 in sheet.iter_rows(values_only=True):
    # if value1[9] == "major":

    for column in sheet.columns:
        excel_text = column[1].value
        if excel_text in states:
            if excel_text in statedict:
                statedict[excel_text] = statedict[excel_text]
            else:
                statedict[excel_text] = 1
    state_names_pandas = pandas.Series(states)
    state_count_pandas = pandas.Series(statedict)
    data = [dict(
        type='choropleth',
        colorscale="BlueRed",
        autocolorscale=False,
        locations=state_names_pandas,
        z=state_count_pandas,
        locationmode='USA-states',
        colorbar=dict(
            title="Wage Data info per state")

    )]
    layout = dict(
        title='How Many Times Each State Has Wage Data Info',
        geo=dict(
            scope='usa',
            projection=dict(type='albers usa'),
            showlakes=True, )
    )
    fig = dict(data=data, layout=layout)
    plotly.offline.plot(fig, filename='bcrainesprint4map.html')

    for value in sheet.iter_rows(values_only=True):
        if value[9] == "major":
            make_wage_database_data(cursor, value[1], value[9], value[8], value[10], value[19], value[7])

        # print(item)
    close_db(conn)

    display_data(get_api_data(), excel_ascend_function())


# except Exception:

# print("delete database before running again!")


if __name__ == '__main__':
    main()
